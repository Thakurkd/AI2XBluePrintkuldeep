import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Load API data
# Since the objective is to strictly use documentation, I mapped the testcases manually 
# based on standard SDET best practices utilizing ONLY the provided doc structures.

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define Header style
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row fill
alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

columns = [
    "Scenario", "TID", "TestCase Description", "PreCondition", "Test Steps",
    "Expected Result", "Actual Result", "Steps to Execute", "Expected Result.1", 
    "Actual Result.1", "Status", "Executed QA Name", "Misc (Comments)", "Priority", "Is Automated"
]

test_cases = {
    "CreateToken": [
        {
            "Scenario": "Create Token - Valid Credentials",
            "TID": "API_TC_001",
            "TestCase Description": "Verify that an auth token is generated successfully when valid credentials are provided.",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /auth with Content-Type: application/json\n2. Provide valid body: username='admin', password='password123'",
            "Expected Result": "HTTP 200; response body contains 'token' string.",
            "Steps to Execute": "1. Open REST client\n2. Set method to POST, URL to https://restful-booker.herokuapp.com/auth\n3. Add Header: Content-Type: application/json\n4. Add valid JSON payload\n5. Click Send",
            "Misc (Comments)": "Happy path for authentication.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Create Token - Missing Username",
            "TID": "API_TC_002",
            "TestCase Description": "Verify API handles requests appropriately when mandatory 'username' is omitted.",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /auth with Content-Type: application/json\n2. Omit 'username' from request body.",
            "Expected Result": "Appropriate HTTP error code (e.g., 400); response body contains error message (undocumented specific error code in apidoc).",
            "Steps to Execute": "1. Open REST client\n2. Set method to POST, URL to https://restful-booker.herokuapp.com/auth\n3. Add Header: Content-Type: application/json\n4. Add JSON payload without username\n5. Click Send",
            "Misc (Comments)": "Negative testing. Expected error is undocumented in apidoc.",
            "Priority": "High",
            "Is Automated": "No"
        },
        {
            "Scenario": "Create Token - Invalid Credentials",
            "TID": "API_TC_003",
            "TestCase Description": "Verify API rejects token creation with incorrect password.",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /auth with Content-Type: application/json\n2. Provide valid username but incorrect password.",
            "Expected Result": "Appropriate error response (e.g., HTTP 401) or HTTP 200 with 'reason: Bad credentials' (undocumented behavior).",
            "Steps to Execute": "1. Open REST client\n2. Set POST https://restful-booker.herokuapp.com/auth\n3. Add Header: Content-Type: application/json\n4. Add JSON payload with wrong password\n5. Click Send",
            "Misc (Comments)": "Negative testing for invalid auth.",
            "Priority": "High",
            "Is Automated": "No"
        }
    ],
    "CreateBooking": [
        {
            "Scenario": "Create Booking - Valid Payload (JSON)",
            "TID": "API_TC_004",
            "TestCase Description": "Verify that a booking is created successfully when a valid JSON request payload is submitted.",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /booking with Content-Type: application/json, Accept: application/json\n2. Provide a valid JSON body containing firstname, lastname, totalprice, depositpaid, bookingdates (checkin, checkout), additionalneeds.",
            "Expected Result": "HTTP 200; response body contains bookingid (integer) and the submitted booking object.",
            "Steps to Execute": "1. Open REST client\n2. Set POST https://restful-booker.herokuapp.com/booking\n3. Set Headers: Content-Type: application/json, Accept: application/json\n4. Paste valid JSON body\n5. Click Send",
            "Misc (Comments)": "Validates JSON structure creation.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Create Booking - Valid Payload (XML)",
            "TID": "API_TC_005",
            "TestCase Description": "Verify booking creation using text/xml Content-Type and Accept headers.",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /booking with Content-Type: text/xml, Accept: application/xml\n2. Provide valid XML body with required fields.",
            "Expected Result": "HTTP 200; response body is XML format containing bookingid and booking data.",
            "Steps to Execute": "1. Open REST client\n2. Set POST https://restful-booker.herokuapp.com/booking\n3. Set Headers: Content-Type: text/xml, Accept: application/xml\n4. Paste valid XML body\n5. Click Send",
            "Misc (Comments)": "Validates Content-Type header parsing for XML.",
            "Priority": "Medium",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Create Booking - Missing Mandatory Field (totalprice)",
            "TID": "API_TC_006",
            "TestCase Description": "Verify API behavior when a mandatory field (totalprice) is missing from the payload.",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /booking with Content-Type: application/json\n2. Send body omitting 'totalprice'.",
            "Expected Result": "Appropriate error code (e.g., HTTP 400); response indicates missing field (undocumented in apidoc).",
            "Steps to Execute": "1. Open REST client\n2. Set POST https://restful-booker.herokuapp.com/booking\n3. Set Header Content-Type: application/json\n4. Paste JSON omitting 'totalprice'\n5. Click Send",
            "Misc (Comments)": "Negative testing. Error response undocumented.",
            "Priority": "High",
            "Is Automated": "No"
        },
        {
            "Scenario": "Create Booking - Invalid Date Format",
            "TID": "API_TC_007",
            "TestCase Description": "Verify API validation for checkin/checkout date format (should be YYYY-MM-DD).",
            "PreCondition": "API server is reachable.",
            "Test Steps": "1. Send POST /booking with Content-Type: application/json\n2. Send body with 'bookingdates.checkin' date format as DD-MM-YYYY.",
            "Expected Result": "Appropriate error code; response indicates invalid date format.",
            "Steps to Execute": "1. Open REST client\n2. Set POST https://restful-booker.herokuapp.com/booking\n3. Set Header Content-Type: application/json\n4. Paste JSON with invalid date format\n5. Click Send",
            "Misc (Comments)": "Boundary/input validation.",
            "Priority": "Medium",
            "Is Automated": "No"
        }
    ],
    "DeleteBooking": [
        {
            "Scenario": "Delete Booking - Valid Token (Cookie)",
            "TID": "API_TC_008",
            "TestCase Description": "Verify booking is deleted successfully using a valid Cookie token.",
            "PreCondition": "A booking exists with a known ID. A valid auth token is generated.",
            "Test Steps": "1. Send DELETE /booking/{id}\n2. Pass Header 'Cookie: token=<valid_token>'",
            "Expected Result": "HTTP 201 Created (as per apidoc); response body contains 'OK'.",
            "Steps to Execute": "1. Open REST client\n2. Set DELETE https://restful-booker.herokuapp.com/booking/{id}\n3. Add Header Cookie: token=<valid_token>\n4. Click Send",
            "Misc (Comments)": "Documented success code is 201.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Delete Booking - Basic Auth",
            "TID": "API_TC_009",
            "TestCase Description": "Verify booking is deleted successfully using Basic Auth header.",
            "PreCondition": "A booking exists with a known ID.",
            "Test Steps": "1. Send DELETE /booking/{id}\n2. Pass Header 'Authorization: Basic YWRtaW46cGFzc3dvcmQxMjM='",
            "Expected Result": "HTTP 201 Created; response body contains 'OK'.",
            "Steps to Execute": "1. Open REST client\n2. Set DELETE https://restful-booker.herokuapp.com/booking/{id}\n3. Add Header Authorization with Basic Auth\n4. Click Send",
            "Misc (Comments)": "Alternative auth method.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Delete Booking - Missing Auth",
            "TID": "API_TC_010",
            "TestCase Description": "Verify API rejects delete request without auth headers.",
            "PreCondition": "A booking exists.",
            "Test Steps": "1. Send DELETE /booking/{id} without Cookie or Authorization header.",
            "Expected Result": "Appropriate error code (e.g., HTTP 403 Forbidden).",
            "Steps to Execute": "1. Open REST client\n2. Set DELETE https://restful-booker.herokuapp.com/booking/{id}\n3. Do not add auth headers\n4. Click Send",
            "Misc (Comments)": "Auth validation.",
            "Priority": "High",
            "Is Automated": "No"
        },
        {
            "Scenario": "Delete Booking - Invalid ID",
            "TID": "API_TC_011",
            "TestCase Description": "Verify API behavior when deleting a non-existent booking ID.",
            "PreCondition": "Valid auth token is available.",
            "Test Steps": "1. Send DELETE /booking/{invalid_id} with valid auth header.",
            "Expected Result": "Appropriate error code (e.g., HTTP 404 Not Found).",
            "Steps to Execute": "1. Open REST client\n2. Set DELETE https://restful-booker.herokuapp.com/booking/999999\n3. Add valid auth header\n4. Click Send",
            "Misc (Comments)": "Negative testing for resource existence.",
            "Priority": "Medium",
            "Is Automated": "No"
        }
    ],
    "GetBooking": [
        {
            "Scenario": "Get Booking - Valid ID (JSON)",
            "TID": "API_TC_012",
            "TestCase Description": "Verify retrieval of booking details by valid ID in JSON format.",
            "PreCondition": "A booking exists with a known ID.",
            "Test Steps": "1. Send GET /booking/{id} with Accept: application/json",
            "Expected Result": "HTTP 200 OK; response contains firstname, lastname, totalprice, depositpaid, bookingdates, additionalneeds.",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking/{id}\n3. Set Header Accept: application/json\n4. Click Send",
            "Misc (Comments)": "Happy path retrieval.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Get Booking - Valid ID (XML)",
            "TID": "API_TC_013",
            "TestCase Description": "Verify retrieval of booking details by valid ID in XML format.",
            "PreCondition": "A booking exists with a known ID.",
            "Test Steps": "1. Send GET /booking/{id} with Accept: application/xml",
            "Expected Result": "HTTP 200 OK; response is an XML block containing booking details.",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking/{id}\n3. Set Header Accept: application/xml\n4. Click Send",
            "Misc (Comments)": "Content negotiation testing.",
            "Priority": "Medium",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Get Booking - Non-existent ID",
            "TID": "API_TC_014",
            "TestCase Description": "Verify response for a booking ID that does not exist.",
            "PreCondition": "API server reachable.",
            "Test Steps": "1. Send GET /booking/{invalid_id}",
            "Expected Result": "Appropriate error code (e.g., HTTP 404 Not Found).",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking/999999\n3. Click Send",
            "Misc (Comments)": "Negative testing for resource retrieval.",
            "Priority": "High",
            "Is Automated": "No"
        }
    ],
    "GetBookingIds": [
        {
            "Scenario": "Get Booking Ids - No Filters",
            "TID": "API_TC_015",
            "TestCase Description": "Verify retrieval of all booking IDs without any query parameters.",
            "PreCondition": "Bookings exist in the system.",
            "Test Steps": "1. Send GET /booking",
            "Expected Result": "HTTP 200 OK; response is an array of objects containing bookingid.",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking\n3. Click Send",
            "Misc (Comments)": "Default behavior.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Get Booking Ids - Filter by Name",
            "TID": "API_TC_016",
            "TestCase Description": "Verify retrieval of booking IDs filtered by firstname and lastname.",
            "PreCondition": "A booking exists with matching firstname/lastname.",
            "Test Steps": "1. Send GET /booking?firstname={name}&lastname={name}",
            "Expected Result": "HTTP 200 OK; response array contains only matching booking IDs.",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking?firstname=sally&lastname=brown\n3. Click Send",
            "Misc (Comments)": "Query parameter testing.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Get Booking Ids - Filter by Date",
            "TID": "API_TC_017",
            "TestCase Description": "Verify retrieval of booking IDs filtered by checkin and checkout dates.",
            "PreCondition": "Bookings exist within the date range.",
            "Test Steps": "1. Send GET /booking?checkin=YYYY-MM-DD&checkout=YYYY-MM-DD",
            "Expected Result": "HTTP 200 OK; response array contains booking IDs matching date criteria.",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking?checkin=2014-03-13&checkout=2014-05-21\n3. Click Send",
            "Misc (Comments)": "Query parameter Date filtering.",
            "Priority": "Medium",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Get Booking Ids - Invalid Filter Parameter",
            "TID": "API_TC_018",
            "TestCase Description": "Verify response when invalid query parameter is provided.",
            "PreCondition": "API server reachable.",
            "Test Steps": "1. Send GET /booking?invalidparam=value",
            "Expected Result": "HTTP 200 OK ignoring unknown parameter, or appropriate error code.",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/booking?invalidparam=value\n3. Click Send",
            "Misc (Comments)": "Negative testing for unknown parameters.",
            "Priority": "Low",
            "Is Automated": "No"
        }
    ],
    "PartialUpdateBooking": [
        {
            "Scenario": "Partial Update - Valid Payload",
            "TID": "API_TC_019",
            "TestCase Description": "Verify partial update of a booking using a valid JSON payload.",
            "PreCondition": "A booking exists with a known ID. Valid auth token is available.",
            "Test Steps": "1. Send PATCH /booking/{id} with Headers Content-Type: application/json, Cookie: token=<token>\n2. Provide partial body (e.g., only firstname and lastname).",
            "Expected Result": "HTTP 200 OK; response body contains updated booking details.",
            "Steps to Execute": "1. Open REST client\n2. Set PATCH https://restful-booker.herokuapp.com/booking/{id}\n3. Add Headers Content-Type: application/json, Accept: application/json, Cookie: token=abc123\n4. Add partial JSON payload\n5. Click Send",
            "Misc (Comments)": "Happy path for PATCH.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Partial Update - Missing Auth",
            "TID": "API_TC_020",
            "TestCase Description": "Verify partial update fails without valid authorization.",
            "PreCondition": "A booking exists.",
            "Test Steps": "1. Send PATCH /booking/{id} without Cookie or Basic auth headers.\n2. Provide valid partial body.",
            "Expected Result": "Appropriate error code (e.g., HTTP 403 Forbidden).",
            "Steps to Execute": "1. Open REST client\n2. Set PATCH https://restful-booker.herokuapp.com/booking/{id}\n3. Omit auth headers\n4. Add partial JSON payload\n5. Click Send",
            "Misc (Comments)": "Auth validation for PATCH.",
            "Priority": "High",
            "Is Automated": "No"
        },
        {
            "Scenario": "Partial Update - Invalid Type",
            "TID": "API_TC_021",
            "TestCase Description": "Verify response when partial update sends invalid data type for a field.",
            "PreCondition": "A booking exists. Valid auth token.",
            "Test Steps": "1. Send PATCH /booking/{id} with valid auth.\n2. Set 'totalprice' to a string instead of a number.",
            "Expected Result": "Appropriate error code (e.g., HTTP 400 Bad Request).",
            "Steps to Execute": "1. Open REST client\n2. Set PATCH https://restful-booker.herokuapp.com/booking/{id}\n3. Add valid auth header\n4. Add JSON payload with 'totalprice': 'string'\n5. Click Send",
            "Misc (Comments)": "Input validation testing.",
            "Priority": "Medium",
            "Is Automated": "No"
        }
    ],
    "UpdateBooking": [
        {
            "Scenario": "Update Booking - Valid Payload",
            "TID": "API_TC_022",
            "TestCase Description": "Verify full update of a booking using a valid JSON payload.",
            "PreCondition": "A booking exists with a known ID. Valid auth token is available.",
            "Test Steps": "1. Send PUT /booking/{id} with Headers Content-Type: application/json, Cookie: token=<token>\n2. Provide full valid body.",
            "Expected Result": "HTTP 200 OK; response body contains fully updated booking details.",
            "Steps to Execute": "1. Open REST client\n2. Set PUT https://restful-booker.herokuapp.com/booking/{id}\n3. Add Headers Content-Type: application/json, Accept: application/json, Cookie: token=abc123\n4. Add full valid JSON payload\n5. Click Send",
            "Misc (Comments)": "Happy path for PUT.",
            "Priority": "High",
            "Is Automated": "Planned"
        },
        {
            "Scenario": "Update Booking - Missing Mandatory",
            "TID": "API_TC_023",
            "TestCase Description": "Verify PUT request fails if mandatory fields are omitted.",
            "PreCondition": "A booking exists. Valid auth token.",
            "Test Steps": "1. Send PUT /booking/{id} with valid auth.\n2. Omit a mandatory field (e.g. firstname).",
            "Expected Result": "Appropriate error code (e.g., HTTP 400 Bad Request).",
            "Steps to Execute": "1. Open REST client\n2. Set PUT https://restful-booker.herokuapp.com/booking/{id}\n3. Add valid auth header\n4. Add JSON payload missing 'firstname'\n5. Click Send",
            "Misc (Comments)": "Negative testing for full update validation.",
            "Priority": "High",
            "Is Automated": "No"
        },
        {
            "Scenario": "Update Booking - Basic Auth",
            "TID": "API_TC_024",
            "TestCase Description": "Verify full update succeeds with Basic auth.",
            "PreCondition": "A booking exists.",
            "Test Steps": "1. Send PUT /booking/{id} with Basic auth header.\n2. Provide full valid body.",
            "Expected Result": "HTTP 200 OK; response body contains fully updated booking details.",
            "Steps to Execute": "1. Open REST client\n2. Set PUT https://restful-booker.herokuapp.com/booking/{id}\n3. Add Header Authorization with Basic Auth\n4. Add full valid JSON payload\n5. Click Send",
            "Misc (Comments)": "Auth validation.",
            "Priority": "Medium",
            "Is Automated": "Planned"
        }
    ],
    "Ping": [
        {
            "Scenario": "Health Check - Valid Request",
            "TID": "API_TC_025",
            "TestCase Description": "Verify that the health check endpoint returns success.",
            "PreCondition": "API server is running.",
            "Test Steps": "1. Send GET /ping",
            "Expected Result": "HTTP 201 Created (as per apidoc).",
            "Steps to Execute": "1. Open REST client\n2. Set GET https://restful-booker.herokuapp.com/ping\n3. Click Send",
            "Misc (Comments)": "Health check.",
            "Priority": "High",
            "Is Automated": "Planned"
        }
    ]
}

for endpoint, cases in test_cases.items():
    ws = wb.create_sheet(title=endpoint)
    
    # Write header
    ws.append(columns)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Write data
    for row_idx, tc in enumerate(cases, 2):
        row = [
            tc.get("Scenario", ""),
            tc.get("TID", ""),
            tc.get("TestCase Description", ""),
            tc.get("PreCondition", ""),
            tc.get("Test Steps", ""),
            tc.get("Expected Result", ""),
            "",  # Actual Result
            tc.get("Steps to Execute", ""),
            tc.get("Expected Result", ""),  # Expected Result.1
            "",  # Actual Result.1
            "",  # Status
            "",  # Executed QA Name
            tc.get("Misc (Comments)", ""),
            tc.get("Priority", ""),
            tc.get("Is Automated", "")
        ]
        ws.append(row)
        
        # Apply formatting
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

# Adjust some specific column widths for better readablity
for ws in wb.worksheets:
    ws.column_dimensions['A'].width = 30 # Scenario
    ws.column_dimensions['B'].width = 15 # TID
    ws.column_dimensions['C'].width = 40 # Description
    ws.column_dimensions['D'].width = 25 # PreCondition
    ws.column_dimensions['E'].width = 45 # Test Steps
    ws.column_dimensions['F'].width = 35 # Expected Result
    ws.column_dimensions['G'].width = 15 # Actual Result
    ws.column_dimensions['H'].width = 45 # Steps to Execute
    ws.column_dimensions['I'].width = 35 # Expected Result
    ws.column_dimensions['J'].width = 15 # Actual Result
    ws.column_dimensions['M'].width = 25 # Misc
    ws.column_dimensions['N'].width = 10 # Priority
    ws.column_dimensions['O'].width = 12 # Is Automated

wb.save("API_TestPlan.xlsx")
print("Excel file generated successfully!")
